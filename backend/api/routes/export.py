from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from api.deps import get_db, get_settings
from config import Settings
from db_models.company import Company
from db_models.export_log import ExportLog
from db_models.scrape_task import ScrapeTask
from schemas.export import ExportRequest, ExportResponse

router = APIRouter(tags=["export"])

EXPORT_COLUMNS = [
    "company_name", "phone", "email", "website", "address",
    "city", "state", "zip", "category", "rating", "review_count", "source",
]


@router.post("/", response_model=ExportResponse)
async def trigger_export(
    body: ExportRequest,
    db: AsyncSession = Depends(get_db),
    settings: Settings = Depends(get_settings),
):
    task_q = await db.execute(select(ScrapeTask).where(ScrapeTask.id == body.task_id))
    if not task_q.scalar_one_or_none():
        raise HTTPException(404, "Task not found")

    q = select(Company).where(Company.task_id == body.task_id, Company.is_duplicate_of.is_(None))
    result = await db.execute(q)
    companies = result.scalars().all()

    exports_dir = Path(settings.exports_dir)
    exports_dir.mkdir(parents=True, exist_ok=True)
    export_id = str(uuid.uuid4())

    if body.format == "csv":
        file_path = str(exports_dir / f"export_{export_id}.csv")
        _export_csv(companies, file_path)
    elif body.format == "json":
        file_path = str(exports_dir / f"export_{export_id}.json")
        _export_json(companies, file_path)
    elif body.format == "excel":
        file_path = str(exports_dir / f"export_{export_id}.xlsx")
        _export_excel(companies, file_path)
    else:
        raise HTTPException(400, f"Unsupported format: {body.format}")

    log = ExportLog(id=export_id, task_id=body.task_id, format=body.format,
                    file_path=file_path, rows_exported=len(companies))
    db.add(log)
    await db.commit()
    await db.refresh(log)
    return log


@router.get("/{export_id}/download")
async def download_export(export_id: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ExportLog).where(ExportLog.id == export_id))
    log = result.scalar_one_or_none()
    if not log or not log.file_path:
        raise HTTPException(404, "Export not found")
    path = Path(log.file_path)
    if not path.exists():
        raise HTTPException(404, "Export file not found on disk")
    media = {"csv": "text/csv", "json": "application/json",
             "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    return FileResponse(path, media_type=media.get(log.format, "application/octet-stream"),
                        filename=path.name)


def _export_csv(companies, file_path: str):
    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(EXPORT_COLUMNS)
        for c in companies:
            writer.writerow([getattr(c, col, None) for col in EXPORT_COLUMNS])


def _export_json(companies, file_path: str):
    data = [{col: getattr(c, col, None) for col in EXPORT_COLUMNS} for c in companies]
    Path(file_path).write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str), encoding="utf-8")


def _export_excel(companies, file_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    for row_idx, c in enumerate(companies, 2):
        for col_idx, col in enumerate(EXPORT_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=getattr(c, col, None))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(file_path)
